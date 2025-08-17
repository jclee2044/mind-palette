import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import csv
from PIL import Image, ImageDraw
import io
from utils import load_database, setup_cross_platform_scrolling, sort_colors_by_rainbow


class AssociationsTab:
    def __init__(self, parent, refresh_callback=None):
        self.parent = parent
        self.refresh_callback = refresh_callback
        self.setup_ui()

    def setup_ui(self):
        for widget in self.parent.winfo_children():
            widget.destroy()

        # Create search frame
        search_frame = tk.Frame(self.parent)
        search_frame.pack(fill="x", padx=10, pady=(10, 5))
        
        tk.Label(search_frame, text="Search:", font=("Arial", 14)).pack(side="left", padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=40)
        self.search_entry.pack(side="left", padx=(0, 5))
        self.search_var.trace("w", self.filter_associations)

        # Load database data
        self.all_associations_data = load_database()

        # Create main frame with scrollbar
        main_frame = tk.Frame(self.parent)
        main_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Create canvas and scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Store the canvas for filtering
        self.associations_canvas = canvas
        self.associations_scrollbar = scrollbar

        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Set up cross-platform scrolling
        setup_cross_platform_scrolling(canvas, self.scrollable_frame)

        # Button frame for side-by-side buttons
        button_frame = tk.Frame(self.parent)
        button_frame.pack(pady=10)
        
        # Export to Excel button (removed Refresh Table button)
        export_button = tk.Button(button_frame, text="Export Data to Excel", command=self.export_associations_to_excel)
        export_button.pack()

        # Initial population of the table
        self.populate_associations_table()

    def refresh_table(self):
        """Refresh the associations table with latest data"""
        self.all_associations_data = load_database()
        self.populate_associations_table(self.search_var.get())

    def populate_associations_table(self, filter_text=""):
        # Clear existing content
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        if not self.all_associations_data:
            tk.Label(self.scrollable_frame, text="No associations found.", font=("Arial", 12)).pack(pady=20)
            return

        # Filter data if search text is provided
        if filter_text:
            filtered_data = []
            filter_lower = filter_text.lower()
            for entry in self.all_associations_data:
                if (filter_lower in entry['xkcd_name'].lower() or 
                    filter_lower in entry['hex'].lower() or 
                    filter_lower in entry['associations'].lower()):
                    filtered_data.append(entry)
            sorted_db = sort_colors_by_rainbow(filtered_data)
        else:
            sorted_db = sort_colors_by_rainbow(self.all_associations_data)

        # Create header
        header_frame = tk.Frame(self.scrollable_frame)
        header_frame.pack(fill="x", pady=(0, 3))
        
        tk.Label(header_frame, text="Color", font=("Arial", 10, "bold"), width=8).pack(side="left", padx=(0, 2))
        tk.Label(header_frame, text="Name", font=("Arial", 10, "bold"), width=8).pack(side="left", padx=(0, 4))
        tk.Label(header_frame, text="Hex Code", font=("Arial", 10, "bold"), width=10).pack(side="left", padx=(15, 2))
        tk.Label(header_frame, text="Associations", font=("Arial", 10, "bold"), width=18).pack(side="left", padx=(0, 2))

        # Create separator
        separator = ttk.Separator(self.scrollable_frame, orient="horizontal")
        separator.pack(fill="x", pady=(0, 3))

        # Create rows for each association
        for entry in sorted_db:
            row_frame = tk.Frame(self.scrollable_frame)
            row_frame.pack(fill="x", pady=0)

            # Color square
            color_canvas = tk.Canvas(row_frame, width=20, height=20, bg=entry['hex'], relief="solid", bd=1)
            color_canvas.pack(side="left", padx=(0, 2))

            # Color name
            name_label = tk.Label(row_frame, text=entry['xkcd_name'], width=12, anchor="w")
            name_label.pack(side="left", padx=(0))

            # Hex code
            hex_label = tk.Label(row_frame, text=entry['hex'], width=7, anchor="w")
            hex_label.pack(side="left", padx=(0))

            # Associations (truncated if too long)
            associations = entry['associations']
            if len(associations) > 65:
                associations = associations[:62] + "..."
            assoc_label = tk.Label(row_frame, text=associations, width=45, anchor="w", justify="left")
            assoc_label.pack(side="left", padx=(0, 2))

            # Edit button
            edit_button = tk.Button(row_frame, text="Edit", command=lambda e=entry: self.edit_association(e))
            edit_button.pack(side="left", padx=(0, 2))

            # ✕ delete label styled as a hyperlink
            delete_label = tk.Label(
                row_frame,
                text="✕",
                fg="grey",
                font=("Arial", 10, "bold"),
                bg=row_frame.cget("bg")
            )
            delete_label.pack(side="left", pady=(2))
            delete_label.bind("<Button-1>", lambda e, entry=entry: self.delete_association(entry))

    def filter_associations(self, *args):
        filter_text = self.search_var.get()
        self.populate_associations_table(filter_text)

    def edit_association(self, entry):
        # Create edit dialog
        dialog = tk.Toplevel(self.parent)
        dialog.title(f"Edit Association - {entry['xkcd_name']}")
        dialog.geometry("500x250")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (250)
        y = (dialog.winfo_screenheight() // 2) - (150)
        dialog.geometry(f"500x250+{x}+{y}")

        # Color display
        color_frame = tk.Frame(dialog)
        color_frame.pack(pady=10)
        
        color_canvas = tk.Canvas(color_frame, width=100, height=60, bg=entry['hex'], relief="solid", bd=2)
        color_canvas.pack(side="left", padx=(0, 10))
        
        info_frame = tk.Frame(color_frame)
        info_frame.pack(side="left")
        
        tk.Label(info_frame, text=f"Name: {entry['xkcd_name']}", font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(info_frame, text=f"Hex: {entry['hex']}", font=("Arial", 10)).pack(anchor="w")

        # Associations text area
        tk.Label(dialog, text="Associations:", font=("Arial", 10, "bold")).pack(pady=(0, 5), anchor="w")
        
        text_frame = tk.Frame(dialog)
        text_frame.pack(fill="both", expand=True, padx=10, pady=(0))
        
        text_widget = tk.Text(text_frame, wrap="word", height=4)
        text_widget.insert("1.0", entry['associations'])
        text_widget.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side="right", fill="y")
        text_widget.config(yscrollcommand=scrollbar.set)

        # Buttons
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=10)
        
        def save_changes():
            new_associations = text_widget.get("1.0", tk.END).strip()
            if new_associations:
                # Update the entry
                entry['associations'] = new_associations
                
                # Update the database
                db = load_database()
                for i, db_entry in enumerate(db):
                    if db_entry['hex'] == entry['hex']:
                        db[i] = entry
                        break
                
                # Save to file
                with open("db/associations.json", "w") as f:
                    json.dump(db, f, indent=4)
                
                dialog.destroy()
                self.setup_ui()  # Refresh the table
            else:
                messagebox.showerror("Error", "Associations cannot be empty.")
        
        def cancel():
            dialog.destroy()
        
        save_button = tk.Button(button_frame, text="Save", command=save_changes)
        save_button.pack(side="left", padx=5)
        
        cancel_button = tk.Button(button_frame, text="Cancel", command=cancel)
        cancel_button.pack(side="left", padx=5)

    def delete_association(self, entry):
        confirm = messagebox.askyesno("Delete", f"Delete association for {entry['xkcd_name']}?")
        if not confirm:
            return

        db = load_database()
        db = [e for e in db if e["hex"] != entry["hex"]]

        with open("db/associations.json", "w") as f:
            json.dump(db, f, indent=4)

        self.all_associations_data = db
        self.populate_associations_table(self.search_var.get())

    def export_associations_to_excel(self):
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            messagebox.showerror("Error", "openpyxl is required for Excel export. Please install it with: pip install openpyxl")
            return

        db = load_database()
        if not db:
            messagebox.showinfo("Export", "No associations to export.")
            return

        # Sort the data by hue (same as table display)
        sorted_db = sort_colors_by_rainbow(db)

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel File"
        )

        if not file_path:
            return  # User canceled

        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Color Associations"

            # Set column headers
            headers = ["Color", "Name", "Hex Code", "Associations"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

            # Add data rows (using sorted data)
            for row, entry in enumerate(sorted_db, 2):
                # Color name
                ws.cell(row=row, column=2, value=entry['xkcd_name'])
                
                # Hex code
                ws.cell(row=row, column=3, value=entry['hex'])
                
                # Associations
                ws.cell(row=row, column=4, value=entry['associations'])

                # Create color square image
                hex_color = entry['hex']
                if hex_color.startswith('#'):
                    # Convert hex to RGB
                    r = int(hex_color[1:3], 16)
                    g = int(hex_color[3:5], 16)
                    b = int(hex_color[5:7], 16)
                    
                    # Create a small image for the color
                    img = Image.new('RGB', (20, 15), (r, g, b))
                    img_io = io.BytesIO()
                    img.save(img_io, format='PNG')
                    img_io.seek(0)
                    
                    # Add image to Excel
                    xl_img = XLImage(img_io)
                    xl_img.width = 20
                    xl_img.height = 15
                    ws.add_image(xl_img, f'A{row}')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Export", f"Associations exported to {file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {str(e)}") 